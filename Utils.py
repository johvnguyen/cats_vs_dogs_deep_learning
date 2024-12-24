import openpyxl
import os

def save_training_stats(experiment_name, losses, accuracies, test_losses, test_accuracies) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet.cell(row = 1, column = 1).value = 'Epoch'
    sheet.cell(row = 1, column = 2).value = 'Training Loss'
    sheet.cell(row = 1, column = 3).value = 'Training Accuracy'
    sheet.cell(row = 1, column = 4).value = 'Test Loss'
    sheet.cell(row = 1, column = 5).value = 'Test Loss'



    for i in range(len(accuracies)):
        sheet.cell(row = i+2, column = 1).value = i + 1
        sheet.cell(row = i+2, column = 2).value = losses[i]
        sheet.cell(row = i+2, column = 3).value = accuracies[i]
        sheet.cell(row = i+2, column = 4).value = test_losses[i]
        sheet.cell(row = i+2, column = 5).value = test_accuracies[i]

    filename = ''.join([experiment_name, '_data.xlsx'])
    workbook.save(os.path.join('experiment_data', filename))
